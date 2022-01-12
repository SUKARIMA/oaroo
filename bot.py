print("==================================")
from logging import error
import re
import discord
import time
import os
import random
import datetime
import openpyxl
oc = ("")

canrec = False



def resource_path(relative_path):
    try:
        base_path = os.system._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

client = discord.Client()

iD = False

@client.event
async def on_ready():
    print("이름 :",client.user.name)
    print("ID :",client.user.id)
    print("======로그인 성공======")
    s = random.randint(1,3)
    if s==1:
        game = discord.Game("아로야 (할말)")
    if s==2:
        game = discord.Game("아로야 (할말)")
    if s==3:
        game = discord.Game("아로야 (할말)")
    await client.change_presence(status=discord.Status.online,activity=game)    
@client.event
async def on_message(message):
    if message.content.startswith("아로야"):
        msg = "g"
        try:
            msg = message.content.split(" ")[1]
        except IndexError:
            await message.channel.send("왜요?")
        if msg == "안녕":
            await message.channel.send("반가워요!")
        elif msg == "하이":
            await message.channel.sen.d("하이하이!!")
        elif msg == "배워":
                    file = openpyxl.load_workbook("아로.xlsx")
                    sheet = file.active
                    learn = message.content.split(" ")
                    imbed = discord.Embed(color=11321034)
                    a=0
                    for i in range(1, 601):
                        plr = str(sheet["C"+str(i)].value).split("(")[0]
                        if str(plr) == str(message.author.name):
                            a += 1
                    if a == 0:
                        imbed.add_field(name="주의!", value="알려준 메세지는 ***모든서버*** 에서 표시됩니다! \n기존 명령어 목록에 알려준 명령어가 생길경우 \n그 명령어가 사라질수 있어요! \n ***주의하세요!***", inline=True)                
                        await message.channel.send("{}".format(message.author.mention),embed=imbed)
                    urls = re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*(),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+',message.content.lower())
                    try:
                        if message.mentions[2 or 3]:
                            await message.channel.send("{}, 멘션은 안돼요!".format(message.author.mention))
                    except IndexError:
                        if urls==True:
                            pr = False
                            await message.channel.send("{}, 링크형식은 안받아요!".format(message.author.mention))
                        else:
                            pr = True
                            for v in oc:
                                if learn[2] == v:
                                    pr = False
                                    break
                            for i in range(1, 601):
                                if sheet["A"+str(i)].value == "-" or sheet["A"+str(i)].value == learn[2] and pr == True:
                                    if sheet["A"+str(i)].value == learn[2]:
                                        pr = False
                                    sheet["A"+str(i)].value = learn[2]
                                    if len(learn) > 4:
                                        sheet["B"+str(i)].value = ""
                                        for v in range(len(learn)-3):
                                            sheet["B"+str(i)].value = sheet["B"+str(i)].value + " " + str(learn[v+3])
                                    else:
                                        sheet["B"+str(i)].value = learn[3]
                                    sheet["c"+str(i)].value = str(message.author.name) + "(" + str(message.author.id) + ")"
                                    #print("valueLoaded")
                                    break
                            if pr:
                                await message.channel.send("`"+str(sheet["A"+str(i)].value)+"` 은`"+str(sheet["B"+str(i)].value)+"` 이라고요? 감사합니다!")
                                file.save("아로.xlsx")
                                print("Saved Successfully")
                            else:
                                await message.channel.send("{}, 이미 아는 거예요!".format(message.author.mention))
                            #
        elif msg == "잊어":
            file = openpyxl.load_workbook("아로.xlsx")
            sheet = file.active
            delete = message.content.split(" ")
            pr = False
            it = False
            for i in range(1, 601):
                if sheet["A"+str(i)].value == delete[2]:
                    plr = sheet["C"+str(i)].value.split("(")[0]
                    if message.author.name == str(plr) or str(message.author.id) == "0":
                        pr = True
                        it = True
                        sheet["A"+str(i)].value = "-"
                        sheet["B"+str(i)].value = ""
                        sheet["c"+str(i)].value = ""
                    else:
                        await message.channel.send("{}, 사용자님이 알려주신 게 아니에요! ".format(message.author.mention))
                        pr = False
                        it = True
                    break
            if pr and it:
                await message.channel.send("어라..? 그게 뭐였죠..?")
                file.save("아로.xlsx")
                print("Saved Successfully")
            elif pr == False and it == False:
                await message.channel.send("{}, 모르는 거예요!".format(message.author.mention))
        else:
            file = openpyxl.load_workbook("아로.xlsx")
            sheet = file.active
            learn = message.content.split(" ")
            sss = False
            for i in range(1, 601):
                if sheet["A"+str(i)].value == msg:
                    sss = await message.channel.send(str(sheet["B"+str(i)].value) +" \n```"+ str(sheet["C"+str(i)].value) + "님이 알려주셨어요!```")
            if sss == False:
                s = random.randint(0,3)
                if s==0:
                    await message.channel.send("그게 뭔가요?")
                elif s==1:
                    await message.channel.send("?")
                elif s==2:
                    await message.channel.send("네?")
                else :
                    await message.channel.send("어... 모르겠어요!")
client.run("OTMwNDI5MTU0MTExOTQ2NzUz.Yd1vrA.JXl1Iy7QaJclj8LwhI9tF8PEppo")